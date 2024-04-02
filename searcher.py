import os
import logging
import sys
from openpyxl import load_workbook

# Suppress openpyxl warnings
logging.getLogger('openpyxl').setLevel(logging.ERROR)

def find_string_in_workbook(directory='.', search_string=''):
    found_info = []
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory, filename)
            workbook = load_workbook(filename=file_path)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and search_string.lower() in str(cell.value).lower():
                            found_info.append((filename, sheet, cell.coordinate))
    return found_info

if __name__ == "__main__":
    if len(sys.argv) == 1:
        print("Usage: python script.py <search_string> <directory>")
        sys.exit(1)
    elif len(sys.argv) == 2:
        search_string = sys.argv[1]
        found_cells = find_string_in_workbook(search_string=search_string)
    else:
        search_string = sys.argv[1]
        directory = sys.argv[2]
        found_cells = find_string_in_workbook(directory=directory, search_string=search_string)

    if found_cells:
        print(f"The search string '{search_string}' was found in the following locations:")
        for cell_info in found_cells:
            print(f"In workbook '{cell_info[0]}' in sheet '{cell_info[1]}' at cell '{cell_info[2]}'")
    else:
        print(f"The search string '{search_string}' was not found in any of the files.")
